from __future__ import annotations

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from ariadne.database import Database
from ariadne.eec import eec_name


def export_device_to_excel(db: Database, device_id: int, output_path: str) -> str:
    device = db.get_device_by_id(device_id)
    entries = db.get_bom_entries(device_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    headers = [
        "Item", "Qty", "Reference", "Part Value", "Package",
        "Mounting", "Manufacturer", "Mfr Order Code",
        "Supplier", "Supplier Code", "EEC Category", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, e in enumerate(entries, 2):
        ws.cell(row=i, column=1, value=e["item_number"])
        ws.cell(row=i, column=2, value=e["quantity"])
        ws.cell(row=i, column=3, value=e["reference_designator"])
        ws.cell(row=i, column=4, value=e["part_value"])
        ws.cell(row=i, column=5, value=e["package"])
        ws.cell(row=i, column=6, value=e["mounting_type"])
        ws.cell(row=i, column=7, value=e["manufacturer"])
        ws.cell(row=i, column=8, value=e["manufacturer_order_code"])
        ws.cell(row=i, column=9, value=e["supplier"])
        ws.cell(row=i, column=10, value=e["supplier_order_code"])
        ws.cell(row=i, column=11, value=eec_name(e["eec_category_id"]) if e["eec_category_id"] else "")
        ws.cell(row=i, column=12, value=e["notes"])

    widths = [6, 5, 40, 18, 25, 9, 18, 22, 12, 18, 22, 15]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    return output_path
