from __future__ import annotations

from pathlib import Path

from ariadne.ai_client import DeepSeekClient
from ariadne.config import AppConfig
from ariadne.database import Database
from ariadne.excel_parser import parse_excel_bom
from ariadne.csv_parser import parse_csv_bom
from ariadne.eec import classify_all
from ariadne.mongo_store import RawDataStore
from ariadne.models import Device, ImportResult
from ariadne.pdf_extractor import extract_text_from_pdf
from ariadne.pdf_parser import parse_pdf_bom_text


class Orchestrator:
    def __init__(self, config: AppConfig):
        self._config = config
        self._db = Database(config.database)
        self._ai = DeepSeekClient(config.deepseek)
        self._raw = RawDataStore(config.mongo)

    def process_file(self, file_path: str, device: Device) -> ImportResult:
        path = Path(file_path)
        ext = path.suffix.lower()
        metadata = {
            "device_brand": device.brand,
            "device_model": device.model_name,
            "manufacturer": device.manufacturer,
        }

        raw_content = self._read_raw(file_path, ext)
        if raw_content is not None:
            raw_id = self._raw.store(
                filename=path.name,
                file_format=ext.lstrip("."),
                content=raw_content,
                metadata=metadata,
            )

        if ext in (".xlsx", ".xls"):
            return self._process_excel(file_path, device)
        elif ext == ".csv":
            return self._process_csv(file_path, device)
        elif ext == ".pdf":
            return self._process_pdf(file_path, device)
        else:
            result = ImportResult()
            result.errors.append(f"Unsupported format: {ext}")
            result.success = False
            return result

    def _read_raw(self, file_path: str, ext: str) -> str | None:
        """Estrae il contenuto grezzo leggibile per l'archivio MongoDB."""
        try:
            if ext in (".xlsx", ".xls"):
                from openpyxl import load_workbook
                wb = load_workbook(file_path, data_only=True)
                lines = []
                for ws in wb.worksheets:
                    lines.append(f"[Sheet: {ws.title}]")
                    for row in ws.iter_rows(values_only=True):
                        lines.append("\t".join("" if c is None else str(c) for c in row))
                return "\n".join(lines)
            elif ext == ".csv":
                return Path(file_path).read_text(encoding="utf-8", errors="replace")
            elif ext == ".pdf":
                return extract_text_from_pdf(file_path)
            else:
                return None
        except Exception as e:
            from ariadne.mongo_store import logger
            logger.warning("Could not read raw content for %s: %s", file_path, e)
            return None


    def _process_excel(self, file_path: str, device: Device) -> ImportResult:
        entries = parse_excel_bom(file_path)
        return self._import_entries(entries, device)

    def _process_csv(self, file_path: str, device: Device) -> ImportResult:
        entries = parse_csv_bom(file_path)
        return self._import_entries(entries, device)

    def _process_pdf(self, file_path: str, device: Device) -> ImportResult:
        result = ImportResult()

        text = extract_text_from_pdf(file_path)
        if not text.strip():
            result.errors.append("No text extracted from PDF")
            result.success = False
            return result

        result.warnings.append(f"Extracted {len(text)} chars from PDF")

        entries = parse_pdf_bom_text(text)
        if entries:
            result.warnings.append(f"Direct parser extracted {len(entries)} components")
            return self._import_entries(entries, device, result)

        if not self._config.deepseek.enabled:
            result.warnings.append(
                "AI extraction is disabled (DEEPSEEK_ENABLED=false). "
                "PDF could not be parsed without AI."
            )
            result.success = False
            return result

        result.warnings.append("Direct parser found no entries, trying AI extraction...")
        try:
            extraction = self._ai.extract_bom(text)
            entries = extraction.entries
            usage = extraction.usage
            result.warnings.append(
                f"AI extracted {len(entries)} components "
                f"(tokens: {usage.total_tokens}, est. cost: ${usage.cost_usd:.5f})"
            )
        except Exception as e:
            result.errors.append(f"AI extraction failed: {e}")
            result.success = False
            return result

        return self._import_entries(entries, device, result)

    def _import_entries(self, entries, device: Device, result: ImportResult | None = None) -> ImportResult:
        if result is None:
            result = ImportResult()
        result.total_rows = len(entries)
        device_id = self._db.find_or_create_device(device)

        for entry in entries:
            try:
                if entry.eec_category_id is None:
                    entry.eec_category_id = classify_all(entry.reference_designator)
                entry_id = self._db.insert_bom_entry(device_id, entry)
                if entry_id is None:
                    result.warnings.append(
                        f"Row {entry.item_number}: duplicate ({entry.reference_designator}), skipped"
                    )
                else:
                    result.imported_rows += 1
            except Exception as e:
                result.failed_rows += 1
                result.errors.append(
                    f"Row {entry.item_number}: {e}"
                )

        result.success = result.failed_rows == 0
        return result

    def get_stats(self) -> dict:
        stats = self._db.get_stats()
        stats["raw_documents"] = self._raw.count()
        stats["raw_available"] = self._raw.available
        return stats

    def close(self):
        self._ai.close()
        self._raw.close()
        self._db.close()
