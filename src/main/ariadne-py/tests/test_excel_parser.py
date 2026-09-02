import tempfile
from pathlib import Path

import openpyxl
import pytest

from ariadne.excel_parser import EXCEL_HEADER_ROW, _detect_mounting_type, parse_excel_bom


def test_detect_mounting_type_smt():
    assert _detect_mounting_type("0603") == "SMT"
    assert _detect_mounting_type("SOT23") == "SMT"
    assert _detect_mounting_type("LQFP48") == "SMT"


def test_detect_mounting_type_tht():
    assert _detect_mounting_type("DIP-8") == "THT"
    assert _detect_mounting_type("SIP-3") == "THT"
    assert _detect_mounting_type("TO-220") == "THT"


def test_detect_mounting_type_unknown():
    assert _detect_mounting_type(None) == "SMT"
    assert _detect_mounting_type("") == "SMT"
    assert _detect_mounting_type("  ") == "SMT"


def _make_workbook(tmp_path: Path) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, EXCEL_HEADER_ROW + 1):
        ws.cell(row=r, column=1, value=f"header{r}")
    ws.cell(row=7, column=1, value=1)
    ws.cell(row=7, column=2, value=14)
    ws.cell(row=7, column=3, value="C1,C5,C7")
    ws.cell(row=7, column=5, value="100nF")
    ws.cell(row=7, column=9, value="0603")
    ws.cell(row=7, column=10, value="KEMET")
    ws.cell(row=7, column=11, value="mfr code")
    ws.cell(row=7, column=12, value="note here")
    ws.cell(row=7, column=13, value="Mouser")
    ws.cell(row=7, column=14, value="1234")

    ws.cell(row=8, column=1, value="Total")
    ws.cell(row=8, column=2, value=99)

    ws.cell(row=9, column=1, value=None)

    ws.cell(row=10, column=1, value=2)
    ws.cell(row=10, column=2, value=3)
    ws.cell(row=10, column=3, value="U1")
    ws.cell(row=10, column=9, value="DIP-8")

    path = tmp_path / "bom.xlsx"
    wb.save(path)
    return str(path)


def test_parse_excel_bom_fields(tmp_path):
    entries = parse_excel_bom(_make_workbook(tmp_path))
    first = entries[0]
    assert first.item_number == 1
    assert first.quantity == 14
    assert first.reference_designator == "C1,C5,C7"
    assert first.part_value == "100nF"
    assert first.package == "0603"
    assert first.manufacturer == "KEMET"
    assert first.manufacturer_order_code == "mfr code"
    assert first.supplier == "Mouser"
    assert first.supplier_order_code == "1234"
    assert first.notes == "note here"
    assert first.mounting_type == "SMT"


def test_parse_excel_bom_skips_non_numeric(tmp_path):
    entries = parse_excel_bom(_make_workbook(tmp_path))
    assert len(entries) == 2
    assert entries[1].reference_designator == "U1"
    assert entries[1].mounting_type == "THT"


def test_parse_excel_bom_missing_file():
    with pytest.raises(Exception):
        parse_excel_bom("/nonexistent/board.xlsx")
