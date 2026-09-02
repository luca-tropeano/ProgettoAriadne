import openpyxl

from ariadne.ai_client import AIExtractionResult, AIUsage
from ariadne.config import AppConfig, DatabaseConfig, DeepSeekConfig
from ariadne.excel_parser import EXCEL_HEADER_ROW
from ariadne.models import BOMEntry, Device
from ariadne.orchestrator import Orchestrator

DEVICE = Device(brand="TEST", model_name="ORCH-TEST", manufacturer="TEST")


class FakeAI:
    def __init__(self):
        self.calls = 0

    def extract_bom(self, text):
        self.calls += 1
        return AIExtractionResult(
            entries=[BOMEntry(item_number=1, quantity=1, reference_designator="R1")],
            usage=AIUsage(prompt_tokens=10, completion_tokens=5, total_tokens=15, cost_usd=0.000004),
        )

    def close(self):
        pass


def _make_orch(monkeypatch, ai_enabled=False):
    config = AppConfig(
        deepseek=DeepSeekConfig(enabled=ai_enabled, api_key="test"),
        database=DatabaseConfig(url="sqlite:///:memory:"),
    )
    orch = Orchestrator(config)
    fake_ai = FakeAI()
    orch._ai = fake_ai
    monkeypatch.setattr(
        "ariadne.orchestrator.extract_text_from_pdf",
        lambda path: "this is not a BOM, no components here",
    )
    return orch, fake_ai


def test_ai_disabled_by_default():
    config = AppConfig()
    assert config.deepseek.enabled is False


def test_ai_enabled_from_env(monkeypatch):
    monkeypatch.setenv("DEEPSEEK_ENABLED", "true")
    assert AppConfig.from_env().deepseek.enabled is True


def test_pdf_without_ai_when_disabled(monkeypatch):
    orch, fake_ai = _make_orch(monkeypatch, ai_enabled=False)
    try:
        result = orch.process_file("dummy.pdf", DEVICE)
    finally:
        orch.close()
    assert fake_ai.calls == 0
    assert result.success is False
    assert any("DEEPSEEK_ENABLED" in w for w in result.warnings)


def test_pdf_ai_used_when_enabled(monkeypatch):
    orch, fake_ai = _make_orch(monkeypatch, ai_enabled=True)
    try:
        result = orch.process_file("dummy.pdf", DEVICE)
    finally:
        orch.close()
    assert fake_ai.calls == 1
    assert result.success is True
    assert result.imported_rows == 1
    assert any("cost" in w for w in result.warnings)


def test_pdf_empty_text_fails(monkeypatch):
    orch, _ = _make_orch(monkeypatch, ai_enabled=False)
    monkeypatch.setattr("ariadne.orchestrator.extract_text_from_pdf", lambda path: "")
    try:
        result = orch.process_file("empty.pdf", DEVICE)
    finally:
        orch.close()
    assert result.success is False
    assert any("No text extracted" in e for e in result.errors)


def test_process_excel_end_to_end(monkeypatch, tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, EXCEL_HEADER_ROW + 1):
        ws.cell(row=r, column=1, value=f"h{r}")
    ws.cell(row=7, column=1, value=1)
    ws.cell(row=7, column=2, value=2)
    ws.cell(row=7, column=3, value="R1,R2")
    ws.cell(row=7, column=5, value="10k")
    ws.cell(row=7, column=9, value="TO-220")
    ws.cell(row=8, column=1, value=2)
    ws.cell(row=8, column=2, value=3)
    ws.cell(row=8, column=3, value="U1")
    xlsx = tmp_path / "bom.xlsx"
    wb.save(xlsx)
    wb.close()

    orch = Orchestrator(
        AppConfig(database=DatabaseConfig(url="sqlite:///:memory:"))
    )
    try:
        result = orch.process_file(str(xlsx), DEVICE)
    finally:
        orch.close()
    assert result.success is True
    assert result.total_rows == 2
    assert result.imported_rows == 2
    assert result.failed_rows == 0


def test_unsupported_format_fails(monkeypatch):
    orch = Orchestrator(AppConfig(database=DatabaseConfig(url="sqlite:///:memory:")))
    try:
        result = orch.process_file("data.bin", DEVICE)
    finally:
        orch.close()
    assert result.success is False
    assert any("Unsupported format" in e for e in result.errors)
